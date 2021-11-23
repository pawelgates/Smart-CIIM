import openpyxl
import os
import easygui

print("*** Select the Delays folder. \n")
delays_folder = easygui.diropenbox(default="G:/My Drive/CIIM test/General Updates/Delays+Cancelled works")

folder_list = os.listdir(delays_folder)
delay_files = []
for id in folder_list:
    if id[0:5] == "Delay":
        delay_files.append(id)

print(f"*** {len(delay_files)} delays were found. \n")

print("*** Select the Weekly Delay file. \n")
weekly_path = easygui.fileopenbox(default="G:/My Drive/CIIM test/General Updates/Delays+Cancelled works")
week_number = weekly_path[-7:-5]

weekly_wb = openpyxl.load_workbook(weekly_path)
weekly_ws = weekly_wb.active

for delay in delay_files:
    empty_row = 0
    i = 4
    while i < 1000:
        if weekly_ws.cell(i, 6).value == None:
            empty_row = i
            break
        i += 1

    delay_wb = openpyxl.load_workbook(delays_folder + "\\" + delay)
    delay_ws = delay_wb.active

    # Copying Data
    weekly_ws.cell(empty_row, 1).value = week_number    # Week Number
    weekly_ws.cell(empty_row, 4).value = delay_ws.cell(3, 2).value    # Date
    weekly_ws.cell(empty_row, 5).value = delay_ws.cell(11, 1).value  # Reason
    weekly_ws.cell(empty_row, 6).value = delay_ws.cell(7, 7).value[:-13]  # Team Leader
    weekly_ws.cell(empty_row, 7).value = delay_ws.cell(8, 4).value  # Section
    weekly_ws.cell(empty_row, 8).value = delay_ws.cell(5, 2).value  # Contract start time
    weekly_ws.cell(empty_row, 9).value = delay_ws.cell(5, 2).value  # TP start time
    weekly_ws.cell(empty_row, 10).value = delay_ws.cell(5, 4).value  # 612 start time
    weekly_ws.cell(empty_row, 11).value = delay_ws.cell(5, 6).value  # Actual start time
    weekly_ws.cell(empty_row, 14).value = delay_ws.cell(6, 2).value  # Contract end time
    weekly_ws.cell(empty_row, 15).value = delay_ws.cell(6, 2).value  # TP end time
    weekly_ws.cell(empty_row, 16).value = delay_ws.cell(6, 4).value  # 612 end time
    weekly_ws.cell(empty_row, 17).value = delay_ws.cell(6, 6).value  # Actual end time
    weekly_ws.cell(empty_row, 21).value = "ISR"
    weekly_ws.cell(empty_row, 22).value = 1     # Foreman
    weekly_ws.cell(empty_row, 23).value = 1     # Team Leaders

    k = 0
    while k < 8:
        if delay_ws.cell(18+k, 1).value == None:
            weekly_ws.cell(empty_row, 24).value = k
            break
        if delay_ws.cell(18+k, 1).value == ".":
            weekly_ws.cell(empty_row, 24).value = 0
            break
        k += 1
    if k == 8:
        weekly_ws.cell(empty_row, 24).value = 8

    if delay_ws.cell(28, 4).value != "No vehicle":
        weekly_ws.cell(empty_row, 25).value = delay_ws.cell(28, 4).value

    # Save changes
    os.chdir(weekly_path[:-28])
    weekly_wb.save(weekly_path[-28:])


