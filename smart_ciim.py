import os
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
import re
import pandas as pd
from pathlib import Path
import shutil
from PIL import ImageTk, Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkcalendar import *


def open_folder():
    global delays_folder_path
    delays_folder_path = filedialog.askdirectory(title="Select the Delays folder")
    global tl_list
    tl_list = []
    global tl_list_internal
    tl_list_internal = os.listdir(delays_folder_path)
    for i in range(len(tl_list_internal)):
        tl_list.append(tl_list_internal[i][:-5])
    print(tl_list)
    tl_listbox.delete(0, END)
    for name in tl_list:
        tl_listbox.insert(END, name)

def open_ciim_folder():
    global fc_ciim_folder_path
    fc_ciim_folder_path = filedialog.askdirectory(title="Select the CIIM folder")

def construction_work_plan_open():
    global construcion_wp_path
    construcion_wp_path = filedialog.askopenfilename(title="Select the Construction Work Plan", filetypes=[("Excel files", "*.xlsx")])
    print(construcion_wp_path)

    construction_wp_workbook = load_workbook(filename=construcion_wp_path)
    construction_wp_worksheet = construction_wp_workbook['Const. Plan']
    maxRow = construction_wp_worksheet.max_row
    maxColumn = construction_wp_worksheet.max_column
    global cp_dates
    cp_dates = []
    for i in range(3, maxRow):
        cell_obj = construction_wp_worksheet.cell(row=i, column=4)
        cell_obj_next = construction_wp_worksheet.cell(row=i+1, column=4)
        if str(cell_obj.value) != str(cell_obj_next.value):
            cp_dates.append(str(cell_obj.value)[:-9])
    cp_dates = list(dict.fromkeys(cp_dates))

def update_combo_list():
    dc_combobox['values'] = cp_dates

def combo_selected(event):
    date_selected = dc_combobox.get() + " 00:00:00"
    global year, month, day, tl_index
    day = date_selected[8:10]
    month = date_selected[5:7]
    year = date_selected[:4]
    print(day)

    team_leaders_list = []
    tl_index = []
    construction_wp_workbook = load_workbook(filename=construcion_wp_path)
    construction_wp_worksheet = construction_wp_workbook['Const. Plan']
    maxRow = construction_wp_worksheet.max_row
    maxColumn = construction_wp_worksheet.max_column
    global cp_dates
    for i in range(3, maxRow):
        cell_obj = construction_wp_worksheet.cell(row=i, column=4)
        if str(cell_obj.value) == date_selected:
            tl_cell = construction_wp_worksheet.cell(row=i, column=13)
            if str(tl_cell.value) != "None":
                temp = str(tl_cell.value)
                temp = re.sub("-|0|1|2|3|4|5|6|7|8|9|\)|\(", "", temp)
                team_leaders_list.append(temp)
                tl_index.append(i)

    dc_tl_listbox.delete(0, END)
    for name in team_leaders_list:
        dc_tl_listbox.insert(END, name)

def go(event):
    global cs
    cs = tl_listbox.curselection()
    tl_name_selected.config(text=tl_listbox.get(cs))
    global team_leader_name
    team_leader_name = tl_listbox.get(cs)
    print(team_leader_name)
    clear_cells()
    load_from_excel()
    line_status()

def dc_tl_selected(event):
    dc_listbox_selection_index = dc_tl_listbox.curselection()
    dc_tl_listbox.itemconfig(dc_listbox_selection_index, bg="#f5d3d0")
    global dc_tl_name, teamLeaderNum
    dc_tl_name = str(dc_tl_listbox.get(dc_listbox_selection_index))
    print(dc_tl_name)
    teamLeaderNum = tl_index[dc_listbox_selection_index[0]]
    print(teamLeaderNum)
    create_delay()

def create_delay():
    temp_cwpw_path = Path(construcion_wp_path)
    temp_cwp_path = temp_cwpw_path.parent
    week_num = str(temp_cwp_path)[-2:]
    print(week_num)

    ciim_folder_path = temp_cwp_path.parent
    dc_delays_f_path = Path(str(ciim_folder_path)+"/General Updates/Delays+Cancelled works/")
    print(dc_delays_f_path)
    if os.path.exists(Path(str(dc_delays_f_path)+"/"+year)) == False:
        os.chdir(dc_delays_f_path)
        os.mkdir(year)
    else:
        print("The year folder already exists")
    if os.path.exists(Path(str(dc_delays_f_path)+"/"+year+"/WW"+week_num)) == False:
        os.chdir(Path(str(dc_delays_f_path)+"/"+year))
        os.mkdir("WW"+week_num)
    else:
        print("The week folder already exists")
    dc_day_folder = day+"."+month+"."+year[:-2]
    print(dc_day_folder)
    if os.path.exists(Path(str(dc_delays_f_path)+"/"+year+"/WW"+week_num+"/"+dc_day_folder)) == False:
        os.chdir(Path(str(dc_delays_f_path)+"/"+year+"/WW"+week_num))
        os.mkdir(dc_day_folder)
    else:
        print("The day folder already exists")
    dc_delays_final_path = Path(str(dc_delays_f_path)+"/"+year+"/WW"+week_num+"/"+dc_day_folder)
    dc_template_path = Path(str(ciim_folder_path)+"/Important doc/Empty reports (templates)")
    dc_template_report_name = "Delay Report template v.01.xlsx"
    dc_new_report_name = "Delay Report "+dc_tl_name+dc_day_folder+".xlsx"
    delay_wb_path = str(dc_delays_final_path) + "/" + dc_new_report_name

    if os.path.exists(Path(delay_wb_path)) == True:
        print("delay already exists")
        status_msg = "Delay Report "+dc_tl_name+dc_day_folder+" already exists!"
        dc_folder_status_label.config(text=status_msg, fg="Red")
    else:
        shutil.copy(Path(str(dc_template_path)+"/"+dc_template_report_name), dc_delays_final_path)
        os.chdir(Path(str(dc_delays_f_path)+"/"+year+"/WW"+week_num+"/"+dc_day_folder))
        os.rename(dc_template_report_name, dc_new_report_name)


        construction_wp_workbook = load_workbook(filename=construcion_wp_path)
        construction_wp_worksheet = construction_wp_workbook['Const. Plan']
        delay_workbook = load_workbook(filename=delay_wb_path)
        delay_worksheet = delay_workbook["Sheet1"]

        delay_worksheet.cell(row=3, column=2).value = dc_day_folder  # Date
        delay_worksheet['B3'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet.cell(row=5, column=6).value = ""  # Start time
        delay_worksheet.cell(row=6, column=6).value = ""  # End time
        delay_worksheet.cell(row=7, column=7).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=13).value  # Team Leader
        delay_worksheet['G7'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet.cell(row=7, column=3).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=11).value  # Foreman
        delay_worksheet['C7'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet.cell(row=5, column=2).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=5).value  # CP Start Time
        delay_worksheet['B5'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet.cell(row=6, column=2).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=6).value  # CP End Time
        delay_worksheet['B6'].fill = PatternFill(bgColor="FFFFFF")

        delay_worksheet.cell(row=32, column=2).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=20).value  # Tofes
        delay_worksheet.cell(row=34, column=2).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=21).value  # WSP
        delay_worksheet.cell(row=33, column=2).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=22).value  # COMMpyDd

        delay_worksheet.cell(row=16, column=1).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=11).value  # Foreman name
        delay_worksheet.cell(row=17, column=1).value = construction_wp_worksheet.cell(row=int(teamLeaderNum), column=13).value  # Team Leader name
        delay_worksheet.cell(row=16, column=5).value = 'Foreman'
        delay_worksheet.cell(row=17, column=5).value = 'Team Leader'
        delay_worksheet.cell(row=16, column=7).value = 'SEMI'
        delay_worksheet.cell(row=17, column=7).value = 'SEMI'

        delay_worksheet.cell(row=8, column=6).value = str(construction_wp_worksheet.cell(row=int(teamLeaderNum), column=7).value) + ' to ' + str(
            construction_wp_worksheet.cell(row=int(teamLeaderNum), column=8).value)
        delay_worksheet['F8'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet.cell(row=8, column=4).value = str(construction_wp_worksheet.cell(row=int(teamLeaderNum), column=9).value) + ' - ' + str(
            construction_wp_worksheet.cell(row=int(teamLeaderNum), column=10).value)
        delay_worksheet['D8'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet.cell(row=8, column=8).value = ""
        delay_worksheet['H8'].fill = PatternFill(bgColor="FFFFFF")

        delay_worksheet.cell(row=28, column=2).value = 'Y'
        delay_worksheet.cell(row=29, column=2).value = 'Y'

        delay_worksheet['B8'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet['F5'].fill = PatternFill(bgColor="FFFFFF")
        delay_worksheet['F6'].fill = PatternFill(bgColor="FFFFFF")

        status_msg = "Delay Report " + dc_tl_name + dc_day_folder + " created!"
        dc_folder_status_label.config(text=status_msg, fg="Green")

        delay_workbook.save(str(delay_wb_path))

def clear_cells():
    frame4_stime_entry.delete(0, END)
    frame4_endtime_entry.delete(0, END)
    frame4_ep_entry.delete(0, END)
    frame4_reason_entry.delete(0, END)
    frame4_w1_entry.delete(0, END)
    frame4_w2_entry.delete(0, END)
    frame4_w3_entry.delete(0, END)
    frame4_w4_entry.delete(0, END)
    frame4_w5_entry.delete(0, END)
    frame4_w6_entry.delete(0, END)
    frame4_w7_entry.delete(0, END)
    frame4_w8_entry.delete(0, END)
    frame4_v1_entry.delete(0, END)
    frame4_workers_var.set(0)
    frame4_vehicles_var.set(0)

def load_from_excel():
    full_file_path = delays_folder_path +"/"+ team_leader_name +".xlsx"
    delay_excel_workbook = load_workbook(filename=full_file_path)
    delay_excel_worksheet = delay_excel_workbook['Sheet1']

    if isinstance(delay_excel_worksheet.cell(row=5, column=6).value, str):
        frame4_stime_entry.insert(0, delay_excel_worksheet.cell(row=5, column=6).value)
    if isinstance(delay_excel_worksheet.cell(row=6, column=6).value, str):
        frame4_endtime_entry.insert(0, delay_excel_worksheet.cell(row=6, column=6).value)
    if isinstance(delay_excel_worksheet.cell(row=8, column=2).value, str):
        frame4_ep_entry.insert(0, delay_excel_worksheet.cell(row=8, column=2).value)
    if isinstance(delay_excel_worksheet.cell(row=11, column=1).value, str):
        frame4_reason_entry.insert(0, delay_excel_worksheet.cell(row=11, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=18, column=1).value, str):
        frame4_w1_entry.insert(0, delay_excel_worksheet.cell(row=18, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=19, column=1).value, str):
        frame4_w2_entry.insert(0, delay_excel_worksheet.cell(row=19, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=20, column=1).value, str):
        frame4_w3_entry.insert(0, delay_excel_worksheet.cell(row=20, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=21, column=1).value, str):
        frame4_w4_entry.insert(0, delay_excel_worksheet.cell(row=21, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=22, column=1).value, str):
        frame4_w5_entry.insert(0, delay_excel_worksheet.cell(row=22, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=23, column=1).value, str):
        frame4_w6_entry.insert(0, delay_excel_worksheet.cell(row=23, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=24, column=1).value, str):
        frame4_w7_entry.insert(0, delay_excel_worksheet.cell(row=24, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=25, column=1).value, str):
        frame4_w8_entry.insert(0, delay_excel_worksheet.cell(row=25, column=1).value)
    if isinstance(delay_excel_worksheet.cell(row=28, column=4).value, str):
        frame4_v1_entry.insert(0, delay_excel_worksheet.cell(row=28, column=4).value)

def save_to_excel():
    full_file_path = delays_folder_path + "/" + team_leader_name + ".xlsx"
    delay_excel_workbook = load_workbook(filename=full_file_path)
    delay_excel_worksheet = delay_excel_workbook['Sheet1']

    delay_excel_worksheet['F5'] = frame4_stime_entry.get()
    delay_excel_worksheet['F6'] = frame4_endtime_entry.get()
    delay_excel_worksheet['B8'] = frame4_ep_entry.get()
    delay_excel_worksheet['A11'] = frame4_reason_entry.get()

    if frame4_workers_var.get() == 0:
        delay_excel_worksheet['A18'] = frame4_w1_entry.get()
        if frame4_w1_entry.get() != "":
            delay_excel_worksheet['E18'] = "Worker"
            delay_excel_worksheet['G18'] = "SEMI"
        delay_excel_worksheet['A19'] = frame4_w2_entry.get()
        if frame4_w2_entry.get() != "":
            delay_excel_worksheet['E19'] = "Worker"
            delay_excel_worksheet['G19'] = "SEMI"
        delay_excel_worksheet['A20'] = frame4_w3_entry.get()
        if frame4_w3_entry.get() != "":
            delay_excel_worksheet['E20'] = "Worker"
            delay_excel_worksheet['G20'] = "SEMI"
        delay_excel_worksheet['A21'] = frame4_w4_entry.get()
        if frame4_w4_entry.get() != "":
            delay_excel_worksheet['E21'] = "Worker"
            delay_excel_worksheet['G21'] = "SEMI"
        delay_excel_worksheet['A22'] = frame4_w5_entry.get()
        if frame4_w5_entry.get() != "":
            delay_excel_worksheet['E22'] = "Worker"
            delay_excel_worksheet['G22'] = "SEMI"
        delay_excel_worksheet['A23'] = frame4_w6_entry.get()
        if frame4_w6_entry.get() != "":
            delay_excel_worksheet['E23'] = "Worker"
            delay_excel_worksheet['G23'] = "SEMI"
        delay_excel_worksheet['A24'] = frame4_w7_entry.get()
        if frame4_w7_entry.get() != "":
            delay_excel_worksheet['E24'] = "Worker"
            delay_excel_worksheet['G24'] = "SEMI"
        delay_excel_worksheet['A25'] = frame4_w8_entry.get()
        if frame4_w8_entry.get() != "":
            delay_excel_worksheet['E25'] = "Worker"
            delay_excel_worksheet['G25'] = "SEMI"
    else:
        delay_excel_worksheet['A18'] = "."

    if frame4_v1_entry.get() == "" and frame4_vehicles_var.get() == 0:
        delay_excel_worksheet['D28'] = frame4_v1_entry.get()
        global vehicle1_var
        vehicle1_var = 0
    else:
        delay_excel_worksheet['D28'] = "No vehicle"
        vehicle1_var = 1

    delay_excel_workbook.save(str(full_file_path))
    clear_cells()
    load_from_excel()
    line_status()

def status_check():
    global status_color

    if start_time == 1 and end_time == 1 and ep_var == 1 and reason_var == 1 and worker1_var == 1 and vehicle1_var == 1:
        frame3_status.config(text="Completed", fg="Green")
        status_color = 1
    else:
        frame3_status.config(text="Not completed", fg="Red")
        status_color = 0

def line_status():
    if frame4_stime_entry.get() == "":
        frame4_stime_entry.config(bg="#f5d3d0")
        global start_time
        start_time = 0
    else:
        frame4_stime_entry.config(bg="#d7f7d8")
        start_time = 1
    if frame4_endtime_entry.get() == "":
        frame4_endtime_entry.config(bg="#f5d3d0")
        global end_time
        end_time = 0
    else:
        frame4_endtime_entry.config(bg="#d7f7d8")
        end_time = 1
    if frame4_ep_entry.get() == "":
        frame4_ep_entry.config(bg="#f5d3d0")
        global ep_var
        ep_var = 0
    else:
        frame4_ep_entry.config(bg="#d7f7d8")
        ep_var = 1
    if frame4_reason_entry.get() == "":
        frame4_reason_entry.config(bg="#f5d3d0")
        global reason_var
        reason_var = 0
    else:
        frame4_reason_entry.config(bg="#d7f7d8")
        reason_var = 1
    if frame4_w1_entry.get() == "" and frame4_workers_var.get() == 0:
        frame4_w1_entry.config(bg="#f5d3d0")
        frame4_w2_entry.config(bg="#f5d3d0")
        frame4_w3_entry.config(bg="#f5d3d0")
        frame4_w4_entry.config(bg="#f5d3d0")
        frame4_w5_entry.config(bg="#f5d3d0")
        frame4_w6_entry.config(bg="#f5d3d0")
        frame4_w7_entry.config(bg="#f5d3d0")
        frame4_w8_entry.config(bg="#f5d3d0")
        global worker1_var
        worker1_var = 0
    else:
        frame4_w1_entry.config(bg="#d7f7d8")
        frame4_w2_entry.config(bg="#d7f7d8")
        frame4_w3_entry.config(bg="#d7f7d8")
        frame4_w4_entry.config(bg="#d7f7d8")
        frame4_w5_entry.config(bg="#d7f7d8")
        frame4_w6_entry.config(bg="#d7f7d8")
        frame4_w7_entry.config(bg="#d7f7d8")
        frame4_w8_entry.config(bg="#d7f7d8")
        worker1_var = 1
    if frame4_v1_entry.get() == "" and frame4_vehicles_var.get() == 0:
        frame4_v1_entry.config(bg="#f5d3d0")
        global vehicle1_var
        vehicle1_var = 0
    else:
        frame4_v1_entry.config(bg="#d7f7d8", text="No vehicle")
        vehicle1_var = 1
    status_check()

def show_frame(frame):
    frame.tkraise()

def pick_date(event):
    global c_day, c_month, c_year, c_date, c_week
    selected_date = calendar.get_date()
    c_day = selected_date[8:10]
    c_month = selected_date[5:7]
    c_year = selected_date[0:4]
    c_date = c_day+"-"+c_month+"-"+c_year
    fc_picked_date_label.config(text=c_date, font=('Helvetica', 9, 'bold'))
    df = pd.Timestamp(selected_date)
    if df.dayofweek == 6:
        c_week = df.weekofyear+1
    else:
        c_week = df.weekofyear
    fc_week_label.config(text=c_week, font=('Helvetica', 9, 'bold'))
    if fc_ciim_folder_path == "":
        fc_status_week.config(text="Choose the CIIM folder", fg="Red", font=('Helvetica', 10, 'bold') )
    else:
        if os.path.exists(fc_ciim_folder_path+"/Working Week N"+str(c_week)) == True:
            fc_status_week.config(text="Working Week N"+str(c_week)+" folder already exists", fg="Green", font=('Helvetica', 10, 'bold'))
        else:
            fc_status_week.config(text="Working Week N"+str(c_week)+" folder doesn't exist", fg="Red", font=('Helvetica', 10, 'bold'))
        if os.path.exists(fc_ciim_folder_path+"/Working Week N"+str(c_week)+"/"+str(c_year)[-2:]+str(c_month)+str(c_day)) == True:
            fc_status_day.config(text=str(c_year)[-2:] + str(c_month) + str(c_day) + " folder already exists", fg="Green", font=('Helvetica', 10, 'bold'))
            fc_ocs_entry.config(state="disabled")
            fc_scada_entry.config(state="disabled")
            fc_ba_entry.config(state="disabled")
            fc_custom_x_label.config(state="disabled")
            fc_custom_entry.config(state="disabled")
            fc_create_BTN.config(state="disabled")
        else:
            fc_status_day.config(text=str(c_year)[-2:] + str(c_month) + str(c_day) + " folder doesn't exist", fg="Red", font=('Helvetica', 10, 'bold'))
            fc_ocs_entry.config(state="normal")
            fc_scada_entry.config(state="normal")
            fc_ba_entry.config(state="normal")
            fc_custom_x_label.config(state="normal")
            fc_custom_entry.config(state="normal")
            fc_create_BTN.config(state="normal")



def create_folders():
    if os.path.exists(fc_ciim_folder_path + "/Working Week N" + str(c_week)) == False:
        os.chdir(fc_ciim_folder_path)
        os.mkdir("Working Week N" + str(c_week))
        fc_status_week.config(text="Working Week N" + str(c_week) + " folder was created", fg="Green",
                              font=('Helvetica', 10, 'bold'))
    if os.path.exists(fc_ciim_folder_path + "/Working Week N" + str(c_week) + "/" + str(c_year)[-2:] + str(c_month) + str(c_day)) == False:
        os.chdir(fc_ciim_folder_path + "/Working Week N" + str(c_week))
        os.mkdir(str(c_year)[-2:] + str(c_month) + str(c_day))
    fc_ciim_report_name = "CIIM Report Table "+str(c_day)+"."+str(c_month)+"."+str(c_year)[-2:]+".xlsx"
    fc_ciim_report_path = fc_ciim_folder_path + "/Working Week N" + str(c_week) + "/" + str(c_year)[-2:] + str(c_month) + str(c_day)
    fc_ciim_template_path = fc_ciim_folder_path + "/Important doc/Empty reports (templates)/CIIM Report Table v.1.xlsx"
    # Copy template
    shutil.copy(fc_ciim_template_path, fc_ciim_report_path)
    os.chdir(fc_ciim_folder_path + "/Working Week N" + str(c_week) + "/" + str(c_year)[-2:] + str(c_month) + str(c_day))
    os.rename("CIIM Report Table v.1.xlsx", fc_ciim_report_name)
    # Create folders
    fc_folder_date_temp = fc_ciim_folder_path + "/Working Week N" + str(c_week) + "/" + str(c_year)[-2:] + str(c_month) + str(c_day)
    os.chdir(fc_folder_date_temp)
    if fc_ocs_entry.get() != "" and fc_ocs_entry.get() != "0":
        for i in range(0, int(fc_ocs_entry.get())):
            os.mkdir("W"+str(i+1))
            os.chdir(fc_folder_date_temp+"/"+"W"+str(i+1))
            os.mkdir("Pictures")
            os.mkdir("Worklogs")
            os.chdir(fc_folder_date_temp)
    os.chdir(fc_folder_date_temp)
    if fc_scada_entry.get() != "" and fc_scada_entry.get() != "0":
        for i in range(0, int(fc_scada_entry.get())):
            os.mkdir("S"+str(i+1))
            os.chdir(fc_folder_date_temp + "/" + "S" + str(i + 1))
            os.mkdir("Pictures")
            os.mkdir("Worklogs")
            os.chdir(fc_folder_date_temp)
    os.chdir(fc_folder_date_temp)
    if fc_ba_entry.get() != "" and fc_ba_entry.get() != "0":
        for i in range(0, int(fc_ba_entry.get())):
            os.mkdir("BA"+str(i+1))
            os.chdir(fc_folder_date_temp + "/" + "BA" + str(i + 1))
            os.mkdir("Pictures")
            os.mkdir("Worklogs")
            os.chdir(fc_folder_date_temp)
    os.chdir(fc_folder_date_temp)
    if fc_custom_x_label.get() != "" and fc_custom_entry.get() != "0" and fc_custom_entry.get() != "":
        for i in range(0, int(fc_custom_entry.get())):
            os.mkdir(fc_custom_x_label.get()+str(i+1))
            os.chdir(fc_folder_date_temp + "/" + fc_custom_x_label.get()+str(i+1))
            os.mkdir("Pictures")
            os.mkdir("Worklogs")
            os.chdir(fc_folder_date_temp)
    os.chdir(fc_folder_date_temp)
    os.mkdir("Foreman")
    os.mkdir("612 Forms")
    os.mkdir("Track possession")
    os.mkdir("TS Worklogs")

    fc_ocs_entry.delete(0, END)
    fc_ocs_entry.config(state="disabled")
    fc_scada_entry.delete(0, END)
    fc_scada_entry.config(state="disabled")
    fc_ba_entry.delete(0, END)
    fc_ba_entry.config(state="disabled")
    fc_custom_x_label.delete(0, END)
    fc_custom_x_label.config(state="disabled")
    fc_custom_entry.delete(0, END)
    fc_custom_entry.config(state="disabled")
    fc_create_BTN.config(state="disabled")

    fc_status_day.config(text=str(c_year)[-2:] + str(c_month) + str(c_day) + " folder was created", fg="Green",
                         font=('Helvetica', 10, 'bold'))

# Root config
root = Tk()
root.resizable(0, 0)
root.geometry("870x496")
root.title("Smart CIIM")

main_menu = Menu(root, background='lightblue', foreground='black', activebackground='#004c99', activeforeground='white')
root.config(menu=main_menu,)

main_menu.add_command(label="Delays Creator", command=lambda: show_frame(delays_creator_frame))
main_menu.add_command(label="Delays Manager", command=lambda: show_frame(delays_manager_frame), )
main_menu.add_command(label="Folders Creator", command=lambda: show_frame(folders_creator_frame))


# Variables
tl_list = []
tl_list_internal = []
delays_folder_path = StringVar()
construcion_wp_path = StringVar()
team_leader_name = StringVar()
cp_dates = []
status_color = IntVar()
start_time = 0
end_time = 0
ep_var = 0
reason_var = 0
worker1_var = 0
vehicle1_var = 0
year = StringVar()
month = StringVar()
day = StringVar()
dc_tl_name = StringVar()
tl_index = []
teamLeaderNum = ""
username = ""
fc_ciim_folder_path = ""
c_day = ""
c_month = ""
c_year = ""
c_week = ""
c_date = ""

# Frames
root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)
delays_creator_frame = Frame(root)
delays_manager_frame = Frame(root)
folders_creator_frame = Frame(root)
main_menu_frame = Frame(root)
show_frame(main_menu_frame)
for frame in (delays_creator_frame, delays_manager_frame, folders_creator_frame, main_menu_frame):
    frame.grid(row=0, column=0, sticky="nsew")

# Main menu Frame

bg_image = ImageTk.PhotoImage(Image.open("IMG_0271.PNG"))
bg_label = Label(main_menu_frame, image=bg_image)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)


# Delay Creator
# Frame1 - File select
dc_frame1 = LabelFrame(delays_creator_frame, text="Construction Work Plan", padx=5, pady=5)
dc_frame1.grid(row=0, column=0, sticky="we")
dc_folderBTN_label = Label(dc_frame1, text="Choose the Construction Work Plan:  ", height=2)
dc_folderBTN_label.grid(row=0, column=0)
dc_folderBTN = Button(dc_frame1, text="Open", command=construction_work_plan_open, width=16)
dc_folderBTN.grid(row=0, column=1, sticky="e")

# Frame2 - Date select
dc_frame2 = LabelFrame(delays_creator_frame, text="Select date", padx=5, pady=5)
dc_frame2.grid(row=1, column=0, sticky="we")
dc_select_date_label = Label(dc_frame2, text="Select date:  ", height=2)
dc_select_date_label.grid(row=0, column=0)
dc_combobox = ttk.Combobox(dc_frame2, values=cp_dates, postcommand=update_combo_list)
dc_combobox.set("Date")
dc_combobox.bind("<<ComboboxSelected>>", combo_selected)
dc_combobox.grid(row=0, column=1)

# Frame3 - Team Leaders checkbox
dc_frame3 = LabelFrame(delays_creator_frame, text="Select team", padx=5, pady=5)
dc_frame3.grid(row=2, column=0, sticky="we")
dc_tl_listbox = Listbox(dc_frame3, width=60, height=21)
dc_tl_listbox.grid(sticky="s")
dc_tl_listbox.bind('<Double-1>', dc_tl_selected)

# Frame4 - Status
dc_frame4 = LabelFrame(delays_creator_frame, text="Status", padx=5, pady=5)
dc_frame4.grid(row=0, column=1, sticky="we")
dc_folder_status_label = Label(dc_frame4, text="", height=2, width=67)
dc_folder_status_label.grid(row=0, column=0)

# Folder Creator
# Frame 1 - Calendar
fc_frame1 = LabelFrame(folders_creator_frame, text="Pick the date", padx=5, pady=5)
fc_frame1.grid(row=0, column=0, sticky="we")
fc_frame1.columnconfigure(0, minsize=400)
fc_frame1.rowconfigure(0, minsize=300)
calendar = Calendar(fc_frame1, selectmode="day", firstweekday="sunday", weekenddays=[6, 7], date_pattern="yyyy-mm-dd")
calendar.grid(row=0, column=0, sticky=W+E+N+S)
calendar.bind("<<CalendarSelected>>", pick_date)

# Frame 2 - Info
fc_frame2 = LabelFrame(folders_creator_frame, text="Selected date", padx=5, pady=5)
fc_frame2.grid(row=1, column=0, sticky="we")
fc_ww_label = Label(fc_frame2, text="WW:")
fc_ww_label.grid(row=0, column=0)
fc_week_label = Label(fc_frame2, text="", width=6, anchor=W)
fc_week_label.grid(row=0, column=1)
fc_date_label = Label(fc_frame2, text="Date:")
fc_date_label.grid(row=0, column=2)
fc_picked_date_label = Label(fc_frame2, text="")
fc_picked_date_label.grid(row=0, column=3)

# Frame 3 - Status
fc_frame3 = LabelFrame(folders_creator_frame, text="Status", padx=5, pady=5)
fc_frame3.grid(row=2, column=0, sticky="we")
fc_frame3.rowconfigure(1, minsize=65)
fc_status_week = Label(fc_frame3, text="")
fc_status_week.grid(row=0, column=0, sticky=W)
fc_status_day = Label(fc_frame3, text="")
fc_status_day.grid(row=1, column=0, sticky=W+N)

# Frame 4 - Status
fc_frame4 = LabelFrame(folders_creator_frame, text="CIIM folder", padx=5, pady=5)
fc_frame4.grid(row=0, column=1, sticky=W+N)
fc_frame4.columnconfigure(1, minsize=310)
fc_ciim_folder = Label(fc_frame4, text="Choose the CIIM folder:")
fc_ciim_folder.grid(row=0, column=0, sticky="w")
fc_ciim_folder_BTN = Button(fc_frame4, text="Open", width=12, height=1, command=open_ciim_folder)
fc_ciim_folder_BTN.grid(row=0, column=1, padx=5, pady=5, sticky=W)

# Frame 5 - Create folders
fc_frame5 = LabelFrame(fc_frame4, text="Create folders", padx=5, pady=5)
fc_frame5.grid(row=1, column=0, sticky=W+E, columnspan=2)
fc_frame5.columnconfigure(3, minsize=120)
fc_ocs_label = Label(fc_frame5, text="OCS works:", width=20, anchor=W)
fc_ocs_label.grid(row=0, column=0, sticky=W)
fc_ocs_w_label = Label(fc_frame5, text="W", anchor=E)
fc_ocs_w_label.grid(row=0, column=1, sticky="e")
fc_ocs_entry = Entry(fc_frame5, state="disabled")
fc_ocs_entry.grid(row=0, column=2, sticky=W)
fc_scada_label = Label(fc_frame5, text="SCADA works:", width=20, anchor=W)
fc_scada_label.grid(row=1, column=0, sticky=W)
fc_scada_s_label = Label(fc_frame5, text="S", anchor=E)
fc_scada_s_label.grid(row=1, column=1, sticky="e")
fc_scada_entry = Entry(fc_frame5, state="disabled")
fc_scada_entry.grid(row=1, column=2, sticky=W)
fc_ba_label = Label(fc_frame5, text="Ben Ari works:", width=20, anchor=W)
fc_ba_label.grid(row=3, column=0, sticky=W)
fc_ba_ba_label = Label(fc_frame5, text="BA", anchor=E)
fc_ba_ba_label.grid(row=3, column=1, sticky="e")
fc_ba_entry = Entry(fc_frame5, state="disabled")
fc_ba_entry.grid(row=3, column=2, sticky=W)
fc_custom_label = Label(fc_frame5, text="Custom works:", width=20, anchor=W)
fc_custom_label.grid(row=4, column=0, sticky=W)
fc_custom_x_label = Entry(fc_frame5, width=5, state="disabled")
fc_custom_x_label.grid(row=4, column=1, sticky="e")
fc_custom_entry = Entry(fc_frame5, state="disabled")
fc_custom_entry.grid(row=4, column=2, sticky=W)
fc_create_BTN = Button(fc_frame5, text="Create", height=2, width=12, anchor="center", state="disabled", command=create_folders)
fc_create_BTN.grid(row=1, column=3, rowspan=3, sticky=N+S)

# Delays manager
# Frame 1 - Folder select
frame1 = LabelFrame(delays_manager_frame, text="Delays folder", padx=5, pady=5)
frame1.grid(row=0, column=0, sticky="we")
folderBTN_label = Label(frame1, text="Choose the delays folder:  ", height=2)
folderBTN_label.grid(row=0, column=0)
folderBTN = Button(frame1, text="Open", command=open_folder, width=16)
folderBTN.grid(row=0, column=1, sticky="e")
# Frame 2 - Teams frame
frame2 = LabelFrame(delays_manager_frame, text="Teams", padx=5, pady=5, width=300)
frame2.grid(row=1, column=0, sticky="sw")
tl_listbox = Listbox(frame2, width=45, height=25)
tl_listbox.grid(sticky="s")
tl_listbox.bind('<Double-1>', go)
# Frame 3 - Name + Status
frame3 = LabelFrame(delays_manager_frame, text="Status", padx=5, pady=5)
frame3.grid(row=0, column=1, sticky="w")
tl_name_label = Label(frame3, text="Selected: ", height=2).grid(row=0, column=0, sticky="sw")
tl_name_selected = Label(frame3, text="None", font='Arial 9 bold', height=2, width=50, anchor="w")
tl_name_selected.grid(row=0, column=1)
frame3_status_label = Label(frame3, text="Status: ", height=2).grid(row=0, column=2, sticky="sw")
frame3_status = Label(frame3, text="Not completed", height=2, width=15, anchor="w")
frame3_status.grid(row=0, column=3, sticky="sw")
frame3_status.config(fg="red", font='Arial 9 bold')
# Frame 4 - Manager
frame4 = Frame(delays_manager_frame, padx=5, pady=5)
frame4.grid(row=1, column=1, sticky="nw")
frame4_stime_label = Label(frame4, text="Start time: ").grid(row=0, column=0)
frame4_stime_entry = Entry(frame4)
frame4_stime_entry.grid(row=0, column=1)
frame4_endtime_label = Label(frame4, text="End time: ").grid(row=0, column=2)
frame4_endtime_entry = Entry(frame4)
frame4_endtime_entry.grid(row=0, column=3)
frame4_ep_label = Label(frame4, text="EP: ").grid(row=0, column=4)
frame4_ep_entry = Entry(frame4)
frame4_ep_entry.grid(row=0, column=5)
frame4_reason_label = Label(frame4, text="Reason: ", anchor="w").grid(row=1, column=0, sticky=W)
frame4_reason_entry = Entry(frame4)
frame4_reason_entry.grid(row=1, column=1, columnspan=5, sticky=W + E)
# workers
frame4_space1_label = Label(frame4, text="", anchor="w").grid(row=2, column=0, sticky=W)
frame4_workers_label = Label(frame4, text="Workers: ", anchor="w").grid(row=3, column=0, sticky=W)
frame4_w1_entry = Entry(frame4)
frame4_w1_entry.grid(row=3, column=1, columnspan=2, sticky=W + E)
frame4_w2_entry = Entry(frame4)
frame4_w2_entry.grid(row=4, column=1, columnspan=2, sticky=W + E)
frame4_w3_entry = Entry(frame4)
frame4_w3_entry.grid(row=5, column=1, columnspan=2, sticky=W + E)
frame4_w4_entry = Entry(frame4)
frame4_w4_entry.grid(row=6, column=1, columnspan=2, sticky=W + E)
frame4_w5_entry = Entry(frame4)
frame4_w5_entry.grid(row=7, column=1, columnspan=2, sticky=W + E)
frame4_w6_entry = Entry(frame4)
frame4_w6_entry.grid(row=8, column=1, columnspan=2, sticky=W + E)
frame4_w7_entry = Entry(frame4)
frame4_w7_entry.grid(row=9, column=1, columnspan=2, sticky=W + E)
frame4_w8_entry = Entry(frame4)
frame4_w8_entry.grid(row=10, column=1, columnspan=2, sticky=W + E)
# vehicles
frame4_space2_label = Label(frame4, text="", anchor="w").grid(row=11, column=0, sticky=W)
frame4_vehicles_label = Label(frame4, text="Vehicles: ", anchor="w").grid(row=12, column=0, sticky=W)
frame4_v1_entry = Entry(frame4)
frame4_v1_entry.grid(row=12, column=1, columnspan=2, sticky=W + E)
# check boxes
frame4_workers_var = IntVar()
frame4_workers_cb = Checkbutton(frame4, text="No workers", variable=frame4_workers_var, command=line_status)
frame4_workers_cb.grid(row=3, column=3)
frame4_vehicles_var = IntVar()
frame4_vehicles_cb = Checkbutton(frame4, text="No vehicles", variable=frame4_vehicles_var, command=line_status)
frame4_vehicles_cb.grid(row=12, column=3)
frame4_save_btn = Button(frame4, text="SAVE", width=16, command=save_to_excel)
frame4_save_btn.grid(row=13, column=5)

root.mainloop()
